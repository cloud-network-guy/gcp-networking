import sys
import json
import yaml
import tomli
import tomli_w
import csv
import os
import pathlib

ENCODING = 'utf-8'
SCOPES = ['https://www.googleapis.com/auth/cloud-platform']
SETTINGS_FILE = 'settings.yaml'
PROFILES_FILE = 'profiles.toml'
CALLS_FILE = 'calls.toml'


def get_home_dir() -> str:

    import platform

    my_os = platform.system().lower()
    if my_os.startswith("win"):
        home_dir = os.environ.get("USERPROFILE")
    else:
        home_dir = os.environ.get("HOME")
    return home_dir


def get_docs_dir() -> str:

    import platform

    home_dir = get_home_dir()
    my_os = platform.system().lower()
    docs_dir = "Documents" if my_os.startswith("darwin") or my_os.startswith("win") else ""
    _ = os.path.join(home_dir, docs_dir)
    return _

async def write_to_excel(sheets: dict, file_name: str = "Book1.xlsx", start_row: int = 1):

    from openpyxl import Workbook, utils

    _ = get_docs_dir()
    output_file = os.path.join(_, file_name)

    wb = Workbook()

    for k, v in sheets.items():

        # Create worksheet
        sheet_name = v.get('description', k)
        ws = wb.create_sheet(sheet_name)
        data = v.get('data', [])

        # Skip if the data doesn't have at least one row or first row isn't a dictionary
        if len(data) < 1 or not isinstance(data[0], dict):
            continue

        # Write field names in the first row
        num_columns = 0
        column_widths = {}
        for column_index, column_name in enumerate(data[0].keys()):
            ws.cell(row=start_row, column=column_index + 1).value = column_name
            num_columns += 1
            column_widths[column_index] = len(str(column_name))

        # Write out rows of data
        for row_num in range(len(data)):
            row_data = [str(value) for value in data[row_num].values()]
            ws.append(list(row_data))

            # Keep track of the largest value for each column
            for column_index, entry in enumerate(row_data):
                column_width = len(str(entry)) + 1 if entry else 1
                if column_index in column_widths:
                    if column_width > column_widths[column_index]:
                        column_widths[column_index] = column_width

        for i in range(num_columns):
            ws.column_dimensions[utils.get_column_letter(i + 1)].width = column_widths[i] + 1

    # Save the file
    del wb['Sheet']
    wb.save(filename=output_file)
    print(f"Wrote data to file: {output_file}")


async def read_data_file(file_name: str, file_format: str = None) -> dict:

    if p := pathlib.Path(file_name):
        if not p.is_file():
            open(p, 'a').close()  # Create an empty file
        if p.stat().st_size == 0:
            return {}  # File exists, but is empty
    else:
        raise f"Error occurred while reading '{file_name}'"

    if not file_format:
        file_format = p.suffix.replace('.', '').lower()

    with open(file_name, mode="rb") as fp:
        if file_format == 'yaml':
            return yaml.load(fp, Loader=yaml.FullLoader)
        elif file_format == 'json':
            return json.load(fp)
        elif file_format == 'toml':
            return tomli.load(fp)
        else:
            raise f"unhandled file format '{file_format}'"


async def write_data_file(file_name: str, file_contents: any = None, file_format: str = None) -> None:

    """
    sub_dir = file_name.split('/')[0]
    if not os.path.exists(sub_dir):
        os.makedirs(sub_dir)

    if not file_format:
        file_format = file_name.split('.')[-1].lower()
    """

    p = pathlib.Path(file_name)
    if not file_format:
        file_format = p.suffix.replace('.', '').lower()

    if file_format == 'csv':
        csvfile = open(file_name, 'w', newline='')
        writer = csv.writer(csvfile)
        writer.writerow(file_contents[0].keys())
        [writer.writerow(row.values()) for row in file_contents]
        csvfile.close()
    elif file_format == 'yaml':
            _ = yaml.dump(file_contents)
    elif file_format == 'json':
            _ = json.dumps(file_contents, indent=4)
    elif file_format == 'toml':
        _ = tomli_w.dumps(file_contents)
    else:
        raise f"unhandled file format '{file_format}'"

    with open(file_name, mode="w") as fp:
        fp.write(_)


async def write_file(file_name: str, file_contents: any = None, file_format: str = None) -> None:

    import aiofiles

    if '/' in file_name:
        sub_dir = file_name.split('/')[0]
        if not os.path.exists(sub_dir):
            os.makedirs(sub_dir)

    file_contents = "" if not file_contents else file_contents
    if isinstance(file_contents, bytes):
        file_contents = file_contents.decode(ENCODING)
    async with aiofiles.open(file_name, mode='w', encoding=ENCODING) as fp:
        await fp.write(file_contents)


async def get_settings(settings_file: str = SETTINGS_FILE) -> dict:
    """
    Get all settings from settings file
    """
    _ = await read_data_file(settings_file)
    return _


async def get_profile(settings: dict, profile: str) -> dict:

    if profiles := settings.get('profiles'):
        if _profile := profiles.get(profile):
            return _profile
        else:
            raise f"profile '{profile}' not found in settings file"
    else:
        raise f"no profiles found in settings file"


async def apply_filter(items: list, settings: dict, options: dict = None) -> list:

    # look for regional filter
    #if regions := options.get('regions'):
    regions = [options.get('region')] if 'region' in options else options.get('regions', [])
    items = [item for item in items if item is not None]
    print('regions', regions)
    if len(regions) > 0:
        items = [item for item in items if item.region in regions]
    # look for network string filter
    if _profile := options.get('profile'):
        profile = await get_profile(settings, _profile)
        if network_string := profile.get('network_string'):
            relevant_networks = [item.network_name for item in items if network_string in item.network_name]
        else:
            relevant_networks = [options.get('network')] if 'network' in options else options.get('networks')
        items = [item for item in items if item.network_name in relevant_networks]
    return items


async def get_calls(calls_file: str = CALLS_FILE) -> dict:
    """
    Get all calls from the calls file
    """
    _ = await read_data_file(calls_file)
    return _


async def get_profiles(profiles_file: str = PROFILES_FILE) -> dict:
    """
    Get all profiles from the profiles file
    """
    _ = await read_data_file(profiles_file)
    return _


async def get_platform_info(request: dict) -> dict:

    """
    Get generic information about the VM, Container, or Serverless platform we're running on
    """

    import google.auth
    import platform
    import distro

    distro_info: dict = distro.info()
    server: tuple = request.get('server', ('localhost', 80))
    _ = {
        'distro_id': distro_info.get('id', "UNKNOWN"),
        'distro_version': distro_info.get('version', "UNKNOWN"),
        'platform_machine': platform.machine(),
        'platform_system': platform.system(),
        'python_version': str(sys.version).split()[0],
        'google_auth_version': google.auth.__version__,
        'kernel_version': platform.release(),
        'server_protocol': "HTTP/" + request.get('http_version', "?/?"),
        'server_hostname': server[0],
        'server_port': server[1],
    }
    return _
