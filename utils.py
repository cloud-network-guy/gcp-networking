import json
import yaml
import tomli
import tomli_w
import csv

ENCODING = 'utf-8'
SCOPES = ['https://www.googleapis.com/auth/cloud-platform']
SETTINGS_FILE = 'settings.yaml'
CALLS_FILE = 'calls.toml'


def get_home_dir() -> str:

    from platform import system
    from os import environ

    if my_os := system().lower():
        if my_os.startswith("win"):
            home_dir = environ.get("USERPROFILE")
            separator = "\\Documents\\"
        else:
            home_dir = environ.get("HOME")
            separator = "/Documents/" if my_os.startswith("darwin") else "/"
        return home_dir + separator


async def write_to_excel(sheets: dict, file_name: str = "Book1.xlsx", start_row: int = 1):

    from openpyxl import Workbook, utils

    output_file = f"{get_home_dir()}{file_name}"

    wb = Workbook()

    for k, v in sheets.items():

        # Create worksheet
        sheet_name = v.get('description', k)
        ws = wb.create_sheet(sheet_name)
        data = v.get('data', [])

        # Skip if the data doesn't have at least one row or first row isn't a dictionary
        if len(data) < 1 or not isinstance(data[0], dict):
            continue

        # Write field names in the first row
        num_columns = 0
        column_widths = {}
        for column_index, column_name in enumerate(data[0].keys()):
            ws.cell(row=start_row, column=column_index + 1).value = column_name
            num_columns += 1
            column_widths[column_index] = len(str(column_name))

        # Write out rows of data
        for row_num in range(len(data)):
            row_data = [str(value) for value in data[row_num].values()]
            ws.append(list(row_data))

            # Keep track of the largest value for each column
            for column_index, entry in enumerate(row_data):
                column_width = len(str(entry)) + 1 if entry else 1
                if column_index in column_widths:
                    if column_width > column_widths[column_index]:
                        column_widths[column_index] = column_width

        for i in range(num_columns):
            ws.column_dimensions[utils.get_column_letter(i + 1)].width = column_widths[i] + 1

    # Save the file
    del wb['Sheet']
    wb.save(filename=output_file)
    print(f"Wrote data to file: {output_file}")


async def read_data_file(file_name: str, file_format: str = None) -> dict:

    from pathlib import Path

    if not file_format:
        file_format = file_name.split('.')[-1].lower()

    if p := Path(file_name):
        if not p.is_file():
            raise f"{file_name} not a valid file"
        if p.stat().st_size == 0:
            return {}  # File exists, but is empty

        with open(file_name, mode="rb") as fp:
            match file_format:
                case 'yaml':
                    return yaml.load(fp, Loader=yaml.FullLoader)
                case 'json':
                    return json.load(fp)
                case 'toml':
                    return tomli.load(fp)
                case _:
                    raise f"unhandled file format '{file_format}'"


async def write_data_file(file_name: str, file_contents: any = None, file_format: str = None) -> None:

    from os import path, makedirs

    sub_dir = file_name.split('/')[0]
    if not path.exists(sub_dir):
        makedirs(sub_dir)

    if not file_format:
        file_format = file_name.split('.')[-1].lower()

    match file_format:
        case 'csv':
            csvfile = open(file_name, 'w', newline='')
            writer = csv.writer(csvfile)
            writer.writerow(file_contents[0].keys())
            [writer.writerow(row.values()) for row in file_contents]
            csvfile.close()
            return
        case 'yaml':
            _ = yaml.dump(file_contents)
        case 'json':
            _ = json.dumps(file_contents, indent=4)
        case 'toml':
            _ = tomli_w.dumps(file_contents)
        case _:
            raise f"unhandled file format '{file_format}'"

    with open(file_name, mode="w") as fp:
        fp.write(_)


async def write_file(file_name: str, file_contents: any = None, file_format: str = None) -> None:

    from os import path, makedirs
    import aiofiles

    if '/' in file_name:
        sub_dir = file_name.split('/')[0]
        if not path.exists(sub_dir):
            makedirs(sub_dir)

    file_contents = "" if not file_contents else file_contents
    if isinstance(file_contents, bytes):
        file_contents = file_contents.decode(ENCODING)
    async with aiofiles.open(file_name, mode='w', encoding=ENCODING) as fp:
        await fp.write(file_contents)


async def get_adc_token(quota_project_id: str = None):

    from google.auth import default
    from google.auth.transport.requests import Request

    try:
        credentials, project_id = default(scopes=SCOPES, quota_project_id=quota_project_id)
        _ = Request()
        credentials.refresh(_)
        return credentials.token  # return access token
    except Exception as e:
        raise e


async def read_service_account_key(file: str) -> dict:

    from platform import system
    from google.oauth2 import service_account
    from google.auth.transport.requests import Request

    # If running on Windows, change forward slashes to backslashes
    if system().lower().startswith("win"):
        file = file.replace("/", "\\")

    try:
        with open(file, 'r') as f:
            _ = json.load(f)
            project_id = _.get('project_id')
    except Exception as e:
        raise e

    try:
        credentials = service_account.Credentials.from_service_account_file(file, scopes=SCOPES)
        _ = Request()
        credentials.refresh(_)
        return {'project_id': project_id, 'access_token': credentials.token}
    except Exception as e:
        raise e


async def get_settings(settings_file: str = None) -> dict:

    try:
        settings_file = settings_file if settings_file else SETTINGS_FILE
        if settings := await read_data_file(settings_file):
            return settings
        else:
            raise f"Could not read settings file: '{settings_file}'"
    except Exception as e:
        raise e


async def get_projects(access_token: str, sort_by: str = None) -> list:

    from gcp_operations import make_api_call
    from main import GCPProject

    try:
        url = "https://cloudresourcemanager.googleapis.com/v1/projects"
        _ = await make_api_call(url, access_token)
        #print(_)
        projects = [GCPProject(project_info) for project_info in _]
        if sort_by in ['name', 'id', 'number', 'create_timestamp']:
            projects = sorted(list(projects), key=lambda x: x.get(sort_by), reverse=False)

    except Exception as e:
        raise e

    return projects


async def get_calls(calls_file: str = None) -> dict:

    try:
        calls_file = calls_file if calls_file else CALLS_FILE
        if calls := await read_data_file(calls_file):
            return calls
        else:
            raise f"Could not read settings file: '{calls_file}'"
    except Exception as e:
        raise e


async def get_version(request: dict) -> dict:

    from platform import system, machine, release, version

    try:
        _ = {
            'os': "{} {}".format(system(), release()),
            'cpu': machine(),
            'python_version': str(version).split()[0],
            'server_protocol': "HTTP/" + request.get('http_version', "?/?"),
        }
        return _
    except Exception as e:
        raise e
